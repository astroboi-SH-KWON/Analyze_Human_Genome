from os import listdir
from os.path import isfile, join
import pandas as pd
import openpyxl
from time import clock

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_fa = ".fa"
        self.ext_xlsx = ".xlsx"

    """
    get_excel : read excel file by sheet_name then make data to dict
    :param
        path :
        sheet_name :
    :return
        dict object : { 'col_name' : {0: 'val0' , 1: 'val1' , 2 : 'val2' ...
    """
    def get_excel(self,path,sheet_name):
        return pd.read_excel(path , sheet_name=sheet_name).to_dict()

    """
    get_data_by_col : get data by key(=col_name)
    :return 
        list object : [ ]
    """
    def get_data_by_col(self, tmp_dict,col_name):
        return tmp_dict[col_name].values()

    def read_file_by_line_to_list(self,path):
        tmp_list = []
        with open(path, "r") as f:
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line != '':
                    tmp_list.append(tmp_line)
                else:
                    break
        return tmp_list

    def get_file_list_from_dir(self,path):
        return [f for f in listdir(path) if isfile(join(path,f))]

    """
    make_excel : make dictionary data to excel file
    :param
         path : 
         result_dict : 
    :return
        excel file in path
    """
    def make_excel1(self, param, result_dict):
        path = param[0]
        col_name = param[1]
        max_cnt = param[2]

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="INDEX")
        sheet.cell(row=row, column=2, value=col_name)
        col_header = 3
        for i in range(max_cnt + 1):
            sheet.cell(row=row, column=col_header, value=i)
            col_header = col_header + 1

        for key, val_cnt in result_dict.items():
            row = row + 1
            sheet.cell(row=row, column=1, value=row-1)
            sheet.cell(row=row, column=2, value=key)
            col = 3
            for j in range(max_cnt + 1):
                if j in val_cnt:
                    sheet.cell(row=row, column=col, value=len(val_cnt[j]))
                else:
                    sheet.cell(row=row, column=col, value='0')
                col = col + 1
        workbook.save(filename=path + col_name + "_" + str(clock()) + self.ext_xlsx)

